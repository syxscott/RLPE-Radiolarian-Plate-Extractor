"""Regression tests for the 6 export-module audit bugs found on
2026-08-17.

EXP-1: ``LocalityRecord`` + ``PaleoCoordinateRecord`` declared
       ``paper_id``; converters populate it; xlsx "论文ID" column
       is no longer blank.

EXP-2: DwC-A ``meta.xml`` uses real tab (``\\t``) and real newline
       (``\\n``) for ``fieldsTerminatedBy`` / ``linesTerminatedBy``.
       Previously the literal two-char strings ``\\t`` / ``\\n``
       (backslash + t/n) were written, failing strict GBIF validators.

EXP-3: ``write_dwca_zip`` now raises ``ValueError`` on duplicate
       ``occurrenceID`` instead of silently producing a non-compliant
       archive.

EXP-4: ``export._csv_cell`` (used by the CLI ``--export-csv`` path)
       and ``exporters.analysis._sanitise_csv_cell`` (used by the
       analysis-view ``write_csv``) now share the same formula-injection
       sanitiser.

EXP-5: ``ml._to_ml_record`` defends against ``None`` confidence,
       species, and panel_id (was ``float(None)`` → ``TypeError``).

EXP-6: ``flatten_for_csv`` is idempotent. Calling it twice on the
       same row does not produce ``_md__*`` overflow keys.
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pytest

from rlpe.export import _csv_cell, export_csv, flatten_for_csv
from rlpe.exporters.analysis import _sanitise_csv_cell
from rlpe.exporters.archive import DwCAOptions, write_dwca_zip
from rlpe.exporters.ml import _to_ml_record
from rlpe.exporters.xlsx import write_xlsx
from rlpe.provenance import build_provenance
from rlpe.schema_models import (
    LocalityRecord,
    PaleoCoordinateRecord,
    PanelMetadata,
    PanelRecord,
    PaperMetadataRecord,
    ProvenanceRecord,
    RunOutput,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_provenance() -> ProvenanceRecord:
    return ProvenanceRecord(**build_provenance().to_dict())


def _make_panel(
    *,
    paper_id: str = "p1",
    figure_id: str = "f1",
    panel_id: str | None = "1",
    species: str | None = "Genus speciesum",
    confidence: float = 0.7,
) -> PanelRecord:
    """Build a single-panel RunOutput for archive / ml / xlsx tests."""
    pm = PaperMetadataRecord(
        title="T",
        authors=["A"],
        year=2020,
        doi="10.1/test",
        source="opendataloader",
        confidence=0.8,
    )
    meta = PanelMetadata()
    return PanelRecord(
        paper_id=paper_id,
        figure_id=figure_id,
        panel_id=panel_id,
        species=species,
        panel_path="/tmp/p.png",
        bbox=[0, 0, 100, 100],
        confidence=confidence,
        metadata=meta,
        paper_metadata=pm,
    )


def _make_run(panels: list[PanelRecord]) -> RunOutput:
    return RunOutput(provenance=_make_provenance(), panels=panels)


# ---------------------------------------------------------------------------
# EXP-1: LocalityRecord + PaleoCoordinateRecord declare paper_id
# ---------------------------------------------------------------------------


def test_locality_record_has_paper_id_field() -> None:
    """LocalityRecord must declare paper_id (audit 2026-08-17 EXP-1).

    Previously the schema omitted it and xlsx wrote an empty "论文ID"
    column for every locality row even though the source ``MatchResult``
    clearly carried ``paper_id``.
    """
    fields = set(LocalityRecord.model_fields.keys())
    assert "paper_id" in fields, (
        "LocalityRecord missing paper_id — xlsx '论文ID' column will be blank"
    )


def test_paleo_coordinate_record_has_paper_id_field() -> None:
    """PaleoCoordinateRecord must declare paper_id (EXP-1)."""
    fields = set(PaleoCoordinateRecord.model_fields.keys())
    assert "paper_id" in fields, (
        "PaleoCoordinateRecord missing paper_id — xlsx '论文ID' column will be blank"
    )


def test_locality_record_round_trip_paper_id() -> None:
    """LocalityRecord(...).model_dump() must carry paper_id through."""
    rec = LocalityRecord(
        locality_id="loc1",
        paper_id="p1",
        name="Méouge",
        country="France",
    )
    d = rec.model_dump()
    assert d["paper_id"] == "p1"


def test_paleo_coordinate_record_round_trip_paper_id() -> None:
    """PaleoCoordinateRecord(...).model_dump() must carry paper_id."""
    rec = PaleoCoordinateRecord(
        paleo_coordinate_id="pc1",
        paper_id="p1",
        locality_id="loc1",
    )
    d = rec.model_dump()
    assert d["paper_id"] == "p1"


def test_xlsx_locality_paper_id_populated() -> None:
    """End-to-end: xlsx localities sheet carries paper_id (EXP-1)."""
    panel = _make_panel()
    run = _make_run([panel])
    run = run.model_copy(
        update={
            "localities": [
                LocalityRecord(
                    locality_id="loc1",
                    paper_id="p1",
                    name="Méouge",
                    country="France",
                )
            ]
        }
    )
    blob = write_xlsx(run)  # in-memory bytes
    assert blob is not None
    # Read back via openpyxl
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["localities"]
    # Header is row 1, first data row is row 2.
    header = [c.value for c in ws[1]]
    assert "论文ID" in header
    pid_col = header.index("论文ID")
    assert ws.cell(row=2, column=pid_col + 1).value == "p1", (
        "xlsx localities '论文ID' column empty; LocalityRecord.paper_id not flowing"
    )


def test_xlsx_paleo_paper_id_populated() -> None:
    """End-to-end: xlsx paleo_coordinates sheet carries paper_id (EXP-1)."""
    panel = _make_panel()
    run = _make_run([panel])
    run = run.model_copy(
        update={
            "paleo_coordinates": [
                PaleoCoordinateRecord(
                    paleo_coordinate_id="pc1",
                    paper_id="p1",
                    locality_id="loc1",
                    paleo_latitude=10.0,
                    paleo_longitude=20.0,
                )
            ]
        }
    )
    blob = write_xlsx(run)  # in-memory bytes
    assert blob is not None
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb["paleo_coordinates"]
    header = [c.value for c in ws[1]]
    assert "论文ID" in header
    pid_col = header.index("论文ID")
    assert ws.cell(row=2, column=pid_col + 1).value == "p1", (
        "xlsx paleo_coordinates '论文ID' column empty; PaleoCoordinateRecord.paper_id not flowing"
    )


# ---------------------------------------------------------------------------
# EXP-2: meta.xml uses real tab and real newline
# ---------------------------------------------------------------------------


def test_dwca_meta_xml_uses_real_tab(tmp_path: Path) -> None:
    """``meta.xml`` ``fieldsTerminatedBy`` must be a real tab (EXP-2).

    Previously the literal two-char string ``\\t`` (backslash + t)
    was written, failing strict GBIF validators. We now use
    ``quoteattr`` from ``xml.sax.saxutils`` which serialises a real
    tab as ``&#9;``.
    """
    panel = _make_panel()
    run = _make_run([panel])
    target = tmp_path / "out.zip"
    write_dwca_zip(run, target)

    with zipfile.ZipFile(target) as zf:
        meta_xml = zf.read("meta.xml").decode("utf-8")

    # Real tab attribute: the XML form is "&#9;" (numeric character
    # reference, what quoteattr emits). The literal two-char string
    # ``\\t`` is forbidden.
    assert "\\t" not in meta_xml, (
        f"meta.xml contains literal '\\t'; expected real-tab escape like '&#9;':\n{meta_xml}"
    )
    assert 'fieldsTerminatedBy="&#9;"' in meta_xml, (
        f"meta.xml fieldsTerminatedBy should be a real-tab escape '&#9;':\n{meta_xml}"
    )
    # Same for linesTerminatedBy: real newline must be serialised as
    # '&#10;' (the quoteattr form). The literal '\\n' is forbidden.
    assert "\\n" not in meta_xml, (
        f"meta.xml contains literal '\\n'; expected real-newline escape '&#10;':\n{meta_xml}"
    )
    assert 'linesTerminatedBy="&#10;"' in meta_xml


def test_dwca_default_options_use_real_chars() -> None:
    """DwCAOptions default fields/lines_terminated_by are real chars (EXP-2)."""
    opts = DwCAOptions()
    assert opts.fields_terminated_by == "\t", (
        f"expected real tab, got {opts.fields_terminated_by!r}"
    )
    assert opts.lines_terminated_by == "\n"


# ---------------------------------------------------------------------------
# EXP-3: occurrenceID uniqueness enforced
# ---------------------------------------------------------------------------


def test_dwca_occurrence_id_unique_enforced(tmp_path: Path) -> None:
    """Duplicate occurrenceID must raise ValueError (EXP-3).

    Bypass RunOutput's own dedup by passing a raw ``dict`` payload
    (write_dwca_zip accepts a dict and only RunOutput's __post_init__
    dedup-drops duplicates). The ``_coerce_run_output_from_dict``
    helper does not call __post_init__, so duplicates survive into
    the export dedup check and trigger the expected ValueError.
    """
    p1 = _make_panel(panel_id="1", species="Genus aaa")
    p2 = _make_panel(panel_id="1", species="Genus bbb")
    # ``write_dwca_zip`` accepts a dict (Phase 58 Plan 1.1) and
    # bypasses RunOutput.__post_init__'s dedup loop, so duplicate
    # occurrenceIDs flow through to the EXP-3 dedup check.
    run = {
        "provenance": {"job_id": "j", "source": "test"},
        "panels": [p1.model_dump() if hasattr(p1, "model_dump") else p1 for p in (p1, p2)],
    }
    target = tmp_path / "out.zip"

    with pytest.raises(ValueError) as exc_info:
        write_dwca_zip(run, target)
    msg = str(exc_info.value)
    # The error message must mention duplicate IDs so the operator
    # can find them.
    assert "Duplicate" in msg or "duplicate" in msg
    assert "occurrenceID" in msg


def test_dwca_unique_occurrence_id_succeeds(tmp_path: Path) -> None:
    """Two panels with distinct occ_ids must export cleanly (EXP-3 sanity)."""
    p1 = _make_panel(panel_id="1")
    p2 = _make_panel(panel_id="2")
    run = _make_run([p1, p2])
    target = tmp_path / "out.zip"
    n = write_dwca_zip(run, target)
    assert n == 2


# ---------------------------------------------------------------------------
# EXP-4: shared formula-injection sanitiser between CLI and analysis paths
# ---------------------------------------------------------------------------


def test_cli_csv_sanitises_formula_injection(tmp_path: Path) -> None:
    """CLI ``--export-csv`` path now sanitises formula injection (EXP-4).

    Previously ``export._csv_cell`` did ``str(v)`` and the
    CLI ``--export-csv`` route produced CSV rows with bare
    ``=cmd|'/c calc'!A1`` cells. The fix routes every value through
    the shared ``_sanitise_csv_cell`` (the same one ``analysis.py``
    has always used).
    """
    rows = [
        {
            "paper_id": "p1",
            "species": "=cmd|'/c calc'!A1",
            "confidence": 0.9,
        }
    ]
    target = tmp_path / "out.csv"
    export_csv(rows, target)
    content = target.read_text(encoding="utf-8-sig")
    lines = content.splitlines()
    # Header line then one data line.
    assert len(lines) == 2
    header = lines[0]
    data = lines[1]
    # The species column must be prefixed with a single quote so
    # Excel does not parse it as a formula. We assert the prefix is
    # present (single quote before the ``=``), which is the
    # CWE-1236 contract.
    assert data.startswith("0.9,p1,'="), (
        f"CLI --export-csv did NOT prefix the formula cell; got:\n{content!r}"
    )
    assert "'=cmd|'/c calc'!A1" in data


def test_shared_sanitiser_used_by_both_paths() -> None:
    """``export._csv_cell`` and ``analysis._sanitise_csv_cell`` must
    share behaviour. EXP-4 contract: same prefixes neutralised."""
    # Sanity: both helpers neutralise the dangerous prefixes.
    danger = "=cmd|'/c calc'!A1"
    assert _csv_cell(danger) == _sanitise_csv_cell(danger)
    assert _csv_cell(danger).startswith("'")
    # + prefix
    assert _csv_cell("+1+2").startswith("'")
    # Tab prefix
    assert _csv_cell("\tx").startswith("'")


# ---------------------------------------------------------------------------
# EXP-5: ml.py defends against None
# ---------------------------------------------------------------------------


def test_ml_split_handles_none_confidence() -> None:
    """``_to_ml_record`` must not crash on ``panel.confidence = None``
    (EXP-5). The previous ``float(None)`` raised TypeError."""
    panel = _make_panel(confidence=0.5)
    # Force confidence to None after construction (matches the audit
    # scenario where a dict round-trip leaves the field unset).
    object.__setattr__(panel, "confidence", None)
    rec = _to_ml_record(panel, "train")
    assert rec["confidence"] == 0.0


def test_ml_split_handles_none_species_and_panel_id() -> None:
    """``_to_ml_record`` must coerce None species / panel_id to '' (EXP-5)."""
    panel = _make_panel(panel_id="1", species="Genus x")
    object.__setattr__(panel, "panel_id", None)
    object.__setattr__(panel, "species", None)
    rec = _to_ml_record(panel, "validation")
    assert rec["species"] == ""
    assert rec["panel_id"] == ""


def test_ml_split_normal_panel_unchanged() -> None:
    """A normal panel still flows through unchanged (EXP-5 sanity)."""
    panel = _make_panel(panel_id="1", species="Genus speciesum")
    rec = _to_ml_record(panel, "train")
    assert rec["paper_id"] == "p1"
    assert rec["species"] == "Genus speciesum"
    assert rec["panel_id"] == "1"
    assert rec["confidence"] == 0.7


# ---------------------------------------------------------------------------
# EXP-6: flatten_for_csv is idempotent
# ---------------------------------------------------------------------------


def test_flatten_for_csv_idempotent() -> None:
    """Calling ``flatten_for_csv`` twice produces no NEW ``_md__*``
    overflow keys on the second call (EXP-6). Previously the second
    call pushed primitives into ``_md__<dst>`` as JSON strings,
    producing mixed downstream types.

    Note: the first call may still produce ``_md__<dst>`` overflow keys
    when the top-level primitive already exists (that's pre-existing
    behaviour). The contract being fixed is that the second call MUST
    be a no-op — no new keys, no JSON-encoding of primitives that
    were previously set as native types.
    """
    row = {
        "paper_id": "p1",
        # Top-level primitive already populated.
        "chronostratigraphy_rank": "epoch",
        "metadata": {
            "chronostratigraphy_rank": "epoch",
            "paleodb": {"taxonomy": "x"},
        },
    }
    once = flatten_for_csv(row)
    assert "chronostratigraphy_rank" in once
    # Snapshot the keys of the first pass so we can compare to the
    # second pass.
    once_keys = set(once.keys())
    twice = flatten_for_csv(once)
    twice_keys = set(twice.keys())
    # Second call must not introduce any new keys.
    new_keys = twice_keys - once_keys
    assert new_keys == set(), f"flatten_for_csv is not idempotent; second call added {new_keys!r}"
    # And the value of chronostratigraphy_rank is still the primitive
    # (string "epoch"), not a JSON-encoded blob.
    assert twice["chronostratigraphy_rank"] == "epoch"


def test_flatten_for_csv_second_call_does_not_jsonify_primitives() -> None:
    """EXP-6 — the specific regression: a primitive set on the first
    pass MUST stay a primitive on the second pass (not be re-encoded
    as a JSON string under ``_md__<dst>``).
    """
    row = {
        "paper_id": "p1",
        "chronostratigraphy_rank": "epoch",  # primitive already set
        "metadata": {"chronostratigraphy_rank": "epoch"},
    }
    once = flatten_for_csv(row)
    twice = flatten_for_csv(once)
    # The primitive must not have become a JSON-encoded string.
    assert isinstance(twice["chronostratigraphy_rank"], str)
    assert twice["chronostratigraphy_rank"] == "epoch"
    assert not twice["chronostratigraphy_rank"].startswith('"')


def test_flatten_for_csv_first_call_still_lifts() -> None:
    """EXP-6 must not regress the normal first-call lift (EXP-6 sanity)."""
    row = {
        "paper_id": "p1",
        "metadata": {
            "paleodb": {"taxonomy": "x"},
            "latitude": 35.7,
        },
    }
    flat = flatten_for_csv(row)
    assert "paleodb" in flat
    assert flat["latitude"] == 35.7


# ---------------------------------------------------------------------------
# End-to-end smoke: build a tiny RunOutput, write xlsx + DwC-A, verify
# paper_id populates and meta.xml uses real tab. (audit "Verify" step)
# ---------------------------------------------------------------------------


def test_end_to_end_smoke(tmp_path: Path) -> None:
    """Smoke: xlsx + DwC-A both flow paper_id; DwC-A uses real tab."""
    run = _make_run([_make_panel()])
    run = run.model_copy(
        update={
            "localities": [
                LocalityRecord(
                    locality_id="loc1",
                    paper_id="p1",
                    name="Méouge",
                    country="France",
                )
            ],
            "paleo_coordinates": [
                PaleoCoordinateRecord(
                    paleo_coordinate_id="pc1",
                    paper_id="p1",
                    locality_id="loc1",
                )
            ],
        }
    )

    # xlsx — paper_id flows into both localities and paleo sheets.
    blob = write_xlsx(run)
    assert blob is not None
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(blob))
    for sheet_name, col in (("localities", "论文ID"), ("paleo_coordinates", "论文ID")):
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]
        assert col in header, f"{sheet_name} missing {col}"
        idx = header.index(col) + 1
        assert ws.cell(row=2, column=idx).value == "p1", (
            f"{sheet_name}.{col} empty; paper_id not flowing"
        )

    # DwC-A — meta.xml uses real-tab escape; occurrence.txt uses
    # real tab between columns.
    target = tmp_path / "out.zip"
    n = write_dwca_zip(run, target)
    assert n == 1
    with zipfile.ZipFile(target) as zf:
        meta = zf.read("meta.xml").decode("utf-8")
        occ = zf.read("occurrence.txt").decode("utf-8")
    assert 'fieldsTerminatedBy="&#9;"' in meta
    # Sanity: occurrence.txt actually uses tabs between columns.
    lines = occ.strip().split("\n")
    header_fields = lines[0].split("\t")
    assert "occurrenceID" in header_fields
