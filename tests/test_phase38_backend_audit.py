"""Phase 38 — backend audit regression tests.

The audit (3 parallel subagents) found ~70 bugs across the backend.
This file pins the highest-impact fixes:

  * config.py — type coercion (BUG #67) + range validation (#68, #69, #70)
  * paleodb.py — pooled HTTP opener + close() (BUG PBDB-2, 5)
  * m3_engine.py — _coerce_label handles list values (BUG m3-1)
  * xlsx.py — atomic write via tmp+os.replace (BUG xlsx-4)
  * archive.py — DwC country field (BUG arch-3)
  * ocr.py — paddleocr 2.x / 3.x compatibility (BUG ocr-1)
  * grobid.py — is_available() probe (BUG GROBID-1)
"""

from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))


# ============================================================
# config.py — type coercion + range validation
# ============================================================
def test_config_coerces_string_to_int():
    """Phase 38: YAML / JSON configs often pass int values as strings.
    PipelineConfig must coerce them, not crash later."""
    from rlpe.config import PipelineConfig

    cfg = PipelineConfig(
        pdf_dir=Path("/tmp"),
        work_dir=Path("/tmp"),
        num_workers="4",  # str, not int
        render_dpi="200",  # str, not int
        min_panel_score="0.8",
    )
    assert isinstance(cfg.num_workers, int)
    assert cfg.num_workers == 4
    assert isinstance(cfg.render_dpi, int)
    assert isinstance(cfg.min_panel_score, float)
    assert cfg.min_panel_score == 0.8


def test_config_rejects_out_of_range_render_dpi():
    from rlpe.config import PipelineConfig

    with pytest.raises(ValueError, match="render_dpi"):
        PipelineConfig(
            pdf_dir=Path("/tmp"),
            work_dir=Path("/tmp"),
            render_dpi=10,  # too low
        )
    with pytest.raises(ValueError, match="render_dpi"):
        PipelineConfig(
            pdf_dir=Path("/tmp"),
            work_dir=Path("/tmp"),
            render_dpi=1000,  # too high
        )


def test_config_rejects_out_of_range_min_panel_score():
    from rlpe.config import PipelineConfig

    with pytest.raises(ValueError, match="min_panel_score"):
        PipelineConfig(
            pdf_dir=Path("/tmp"),
            work_dir=Path("/tmp"),
            min_panel_score=2.0,
        )
    with pytest.raises(ValueError, match="min_panel_score"):
        PipelineConfig(
            pdf_dir=Path("/tmp"),
            work_dir=Path("/tmp"),
            min_panel_score=-0.5,
        )


def test_config_rejects_zero_or_negative_num_workers():
    from rlpe.config import PipelineConfig

    with pytest.raises(ValueError, match="num_workers"):
        PipelineConfig(
            pdf_dir=Path("/tmp"),
            work_dir=Path("/tmp"),
            num_workers=0,
        )


def test_config_rejects_non_numeric_inputs():
    from rlpe.config import PipelineConfig

    with pytest.raises(ValueError, match="num_workers"):
        PipelineConfig(
            pdf_dir=Path("/tmp"),
            work_dir=Path("/tmp"),
            num_workers="not_a_number",
        )


def test_config_suggests_typo_for_unknown_extra_keys(caplog):
    """Phase 38: log warning includes Levenshtein-style suggestions
    for unknown extra keys so the user can spot typos."""
    import logging

    from rlpe.config import PipelineConfig

    with caplog.at_level(logging.WARNING):
        PipelineConfig(
            pdf_dir=Path("/tmp"),
            work_dir=Path("/tmp"),
            extra={"minimax_api_key": "x"},  # typo: should be "MiniMax_api_key"
        )
    assert "minimax_api_key" in caplog.text
    # The suggestion should mention the correct key
    assert "MiniMax_api_key" in caplog.text or "did you mean" in caplog.text


# ============================================================
# paleodb.py — pooled HTTP opener + close()
# ============================================================
def test_paleodb_uses_pooled_opener():
    """Phase 38: consecutive PBDB calls share an HTTP opener (no
    fresh TCP connection per call)."""
    from rlpe.paleodb import PaleoDB

    client = PaleoDB(cache_dir=None, offline=True)
    # Lazy — should start as None
    assert client._opener is None
    client.close()
    assert client._opener is None


def test_paleodb_close_clears_opener():
    """Phase 38: close() releases the pooled HTTP opener."""
    from rlpe.paleodb import PaleoDB

    client = PaleoDB(cache_dir=None, offline=True)
    client._opener = "fake_opener"
    client.close()
    assert client._opener is None


# ============================================================
# m3_engine.py — _coerce_label handles list values
# ============================================================
def test_coerce_label_handles_string():
    from rlpe.m3_engine import _coerce_label

    assert _coerce_label("A") == "A"
    assert _coerce_label("  B  ") == "B"
    assert _coerce_label("") is None


def test_coerce_label_handles_list():
    """Phase 38: M3 sometimes returns visible_label as a list
    ['A', 'B']. Old code produced "['A', 'B']" (Python repr)."""
    from rlpe.m3_engine import _coerce_label

    assert _coerce_label(["A", "B"]) == "A, B"
    assert _coerce_label(["X"]) == "X"
    assert _coerce_label([]) is None
    assert _coerce_label(["A", None, "B"]) == "A, B"


def test_coerce_label_handles_none_and_number():
    from rlpe.m3_engine import _coerce_label

    assert _coerce_label(None) is None
    assert _coerce_label(42) == "42"
    assert _coerce_label(3.14) == "3.14"


# ============================================================
# xlsx.py — atomic write
# ============================================================
def test_xlsx_atomic_write_creates_file():
    """Phase 38: xlsx writer must use tmp+os.replace, not direct
    overwrite, so a crash mid-write doesn't leave a corrupt file."""
    from openpyxl import load_workbook

    from rlpe.exporters.xlsx import write_xlsx

    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "out.xlsx"
        write_xlsx(
            run_output={
                "paper_id": "test",
                "panels": [
                    {
                        "paper_id": "test",
                        "panel_id": "P1",
                        "species": "Species foo",
                        "confidence": 0.9,
                        "page_index": 1,
                        "caption_snippet": "test caption",
                    }
                ],
                "geology_contexts": [],
                "localities": [],
                "paleo_coordinates": [],
            },
            path=out,
        )
        assert out.exists(), "xlsx output must be created"
        # Load to confirm it's a valid xlsx
        wb = load_workbook(out, read_only=True)
        assert "panels" in wb.sheetnames


def test_xlsx_atomic_write_cleans_tmp_on_failure(tmp_path, monkeypatch):
    """Phase 38: if openpyxl.save fails, the .tmp file must be cleaned
    up so we don't leak partial files."""
    from rlpe.exporters import xlsx as xlsx_mod

    out = tmp_path / "out.xlsx"

    # Force wb.save to raise
    def boom(self, *a, **k):
        raise RuntimeError("simulated openpyxl failure")

    monkeypatch.setattr(xlsx_mod.Workbook, "save", boom)
    with pytest.raises(RuntimeError, match="simulated"):
        xlsx_mod.write_xlsx(
            run_output={
                "paper_id": "x",
                "panels": [],
                "geology_contexts": [],
                "localities": [],
                "paleo_coordinates": [],
            },
            path=out,
        )
    # The final output must not exist (since save failed)
    assert not out.exists()
    # The .tmp file must have been cleaned up
    tmp = out.with_suffix(out.suffix + ".tmp")
    assert not tmp.exists(), "tmp file must be removed on failure"


# ============================================================
# archive.py — DwC country field
# ============================================================
def test_dwca_country_field_not_empty():
    """Phase 38: archive._occurrence_row was hard-coding country=''
    even when geo.country was set. Fixed: read from geo.country."""
    import inspect

    from rlpe.exporters.archive import _occurrence_row

    # The hard-coded empty string is gone
    src = inspect.getsource(_occurrence_row)
    assert '"country": ""' not in src, (
        "archive.py still hard-codes country='' instead of reading geo.country"
    )
    # And the new line is there
    assert "geo.country" in src


# ============================================================
# ocr.py — paddleocr 2.x / 3.x compatibility
# ============================================================
def test_normalize_paddle_result_2x_format():
    """Phase 38: paddleocr 2.x returns [ [box, (text, conf)], ... ]."""
    from rlpe.ocr import OCRBackend

    result = (
        [
            [[[0, 0], [10, 0], [10, 10], [0, 10]], ("hello", 0.99)],
            [[[20, 20], [30, 20], [30, 30], [20, 30]], ("world", 0.85)],
        ],
        None,
    )
    out = OCRBackend._normalize_paddle_result(result)
    assert len(out) == 2
    assert out[0] == ([[0, 0], [10, 0], [10, 10], [0, 10]], "hello", 0.99)
    assert out[1] == ([[20, 20], [30, 20], [30, 30], [20, 30]], "world", 0.85)


def test_normalize_paddle_result_3x_format():
    """Phase 38: paddleocr 3.x returns a dict with rec_texts etc."""
    from rlpe.ocr import OCRBackend

    result = {
        "rec_texts": ["hello", "world"],
        "rec_scores": [0.99, 0.85],
        "dt_polys": [
            [[0, 0], [10, 0], [10, 10], [0, 10]],
            [[20, 20], [30, 20], [30, 30], [20, 30]],
        ],
    }
    out = OCRBackend._normalize_paddle_result(result)
    assert len(out) == 2
    assert out[0][1] == "hello"
    assert out[0][2] == 0.99
    assert out[1][1] == "world"


def test_normalize_paddle_result_handles_rec_boxes():
    """Phase 38: paddleocr 3.x rec_boxes is a flat 4-int list
    [x1,y1,x2,y2] (4 corner coords), not a 4-point list."""
    from rlpe.ocr import OCRBackend

    result = {
        "rec_texts": ["hi"],
        "rec_scores": [0.9],
        "rec_boxes": [[0, 0, 50, 20]],
    }
    out = OCRBackend._normalize_paddle_result(result)
    assert len(out) == 1
    assert out[0][1] == "hi"
    box = out[0][0]
    # Normalized to 4-point format
    assert len(box) == 4
    assert all(len(p) == 2 for p in box)


def test_normalize_paddle_result_empty():
    from rlpe.ocr import OCRBackend

    assert OCRBackend._normalize_paddle_result(None) == []
    assert OCRBackend._normalize_paddle_result([]) == []
    assert OCRBackend._normalize_paddle_result({}) == []


# ============================================================
# grobid.py — is_available() probe
# ============================================================
def test_grobid_is_available_returns_false_for_dead_server():
    """Phase 38: is_available() must return False for unreachable
    GROBID server (not raise, not return True)."""
    from rlpe.grobid import GrobidClient

    # Port 1 is reserved and never listens — connection refused
    client = GrobidClient(server_url="http://127.0.0.1:1", timeout=1)
    assert client.is_available(probe_timeout=1.0) is False


def test_grobid_is_available_returns_true_for_real_server():
    """Phase 38: is_available() must return True for a live GROBID
    server. We spin up a local HTTP server on a free port."""
    import http.server
    import socketserver
    import threading

    class IsAliveHandler(http.server.BaseHTTPRequestHandler):
        def do_GET(self):
            if self.path == "/api/isalive":
                self.send_response(200)
                self.end_headers()
                self.wfile.write(b"true")
            else:
                self.send_response(404)
                self.end_headers()

        def log_message(self, *args, **kwargs):
            pass

    with socketserver.TCPServer(("127.0.0.1", 0), IsAliveHandler) as httpd:
        port = httpd.server_address[1]
        t = threading.Thread(target=httpd.serve_forever, daemon=True)
        t.start()
        try:
            from rlpe.grobid import GrobidClient

            client = GrobidClient(server_url=f"http://127.0.0.1:{port}", timeout=1)
            assert client.is_available(probe_timeout=2.0) is True
        finally:
            httpd.shutdown()


def test_grobid_is_available_returns_false_for_404():
    """Phase 38: if the server responds 404 to /api/isalive,
    is_available() must return False."""
    import http.server
    import socketserver
    import threading

    class NotFoundHandler(http.server.BaseHTTPRequestHandler):
        def do_GET(self):
            self.send_response(404)
            self.end_headers()

        def log_message(self, *args, **kwargs):
            pass

    with socketserver.TCPServer(("127.0.0.1", 0), NotFoundHandler) as httpd:
        port = httpd.server_address[1]
        t = threading.Thread(target=httpd.serve_forever, daemon=True)
        t.start()
        try:
            from rlpe.grobid import GrobidClient

            client = GrobidClient(server_url=f"http://127.0.0.1:{port}", timeout=1)
            assert client.is_available(probe_timeout=2.0) is False
        finally:
            httpd.shutdown()
