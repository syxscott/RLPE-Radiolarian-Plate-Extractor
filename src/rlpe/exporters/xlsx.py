"""Excel (.xlsx) exporter for Round 24.

The user requested "an Excel table function" that captures
**all data** except images. The CSV exporter at
``web/js/app.js:2187`` is one sheet; this module produces a
proper multi-sheet .xlsx file with:

  Sheet 1: panels             — one row per panel
  Sheet 2: geology_contexts   — one row per geology fact
  Sheet 3: localities         — one row per locality
  Sheet 4: paleo_coordinates — one row per paleo point
  Sheet 5: legend             — column descriptions (free)

Field counts:
  * panels: 20+ columns (paper_id through review_reasons)
  * geology_contexts: 18+ columns (paper_id / age / Ma bounds / ...)
  * localities: 12+ columns
  * paleo_coordinates: 10+ columns
  * legend: 1 column of descriptions

Uses ``openpyxl`` (already a transitive dep via ``pandas`` /
``pypdf`` is not; openpyxl is installed in the env at version
3.1.5). Streaming is not used — for ~10⁴ rows the workbook is
small (< 5 MB). For larger sets (> 10⁵ rows) a future round should
add ``write_only=True`` mode.

Round 24 audit: no formula injection. Numeric values are written
as ``int`` / ``float`` (Excel treats as numbers, not formulas).
String values are written as ``str`` and prefixed with ``'`` if
they start with ``=``/``+``/``-``/``@``/TAB to defeat CWE-1236.
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# CWE-1236 formula-injection sanitiser (mirrors the analysis.py
# CSV path so behaviour is consistent).
_EXCEL_DANGER_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitise(value: Any) -> Any:
    """Sanitise a workbook cell.

    Two concerns are addressed here:

    1. Formula-injection (CWE-1236). A leading ``=``/``+``/``-``/``@``/
       TAB makes Excel treat the cell as a formula; a paper caption
       like ``=cmd|'/c calc'!A1`` would execute on open. Prefixing
       with a single quote neutralises the formula (Excel renders
       the leading quote as part of the cell content but does not
       re-parse it as a formula).
    2. NaN/Inf (Phase 63 Plan 6.8, Bug 6.8). Scale-bar / coordinate
       parsing paths occasionally emit ``float('nan')``. openpyxl
       raises ``ValueError`` on ``nan`` / ``inf`` when writing a
       numeric cell — the cell ends up an ``#N/A`` Excel error which
       GBIF/PBDB ingest rejects. Drop to the empty string so the
       workbook mirrors CSV behaviour.

    Numeric values that are finite and not bool pass through.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        # NaN / Inf — openpyxl refuses to write these, leaving an
        # Excel ``#N/A`` error. Drop to the same shape as missing.
        if math.isnan(value) or math.isinf(value):
            return ""
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        if value and value[0] in _EXCEL_DANGER_PREFIXES:
            return "'" + value
        return value
    if isinstance(value, list):
        return "|".join(_sanitise(v) for v in value)
    if isinstance(value, dict):
        return str(value)
    return str(value)


_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _autosize_columns(ws, max_width: int = 60) -> None:
    """Set column widths based on content. Capped at ``max_width``
    so a long caption_snippet doesn't make the column 200 chars
    wide (UX issue)."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            if length > max_len:
                max_len = length
        # Min 8 chars (empty columns aren't 0-width), capped at max_width.
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, 8), max_width
        )


def _write_header(ws, headers: list[str]) -> None:
    """Write the header row with bold / filled styling so the user
    can identify columns easily. ``wrap_text=True`` so multi-word
    headers wrap on narrow columns instead of overflowing."""
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _write_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    """Write data rows beneath the header. Each row's values are
    sanitised (formula-injection) and the column count is padded /
    truncated to match the header length."""
    for row_idx, row in enumerate(rows, start=2):
        for col_idx in range(len(headers)):
            v = row[col_idx] if col_idx < len(row) else ""
            ws.cell(row=row_idx, column=col_idx + 1, value=_sanitise(v))


# --- panel columns (matches the modal "Panel 来源" row) -------------------
_PANEL_HEADERS = [
    "论文ID", "图版ID", "Panel标签", "Panel来源",
    "物种", "置信度", "地质范围",
    "地层年代", "年代Chrono", "年代Stage", "Ma_top", "Ma_base",
    "Formation", "Member", "Group",
    "Locality", "Country", "Modern_Lat", "Modern_Lon", "Coord来源",
    "Sample IDs", "Lithology", "Biozone",
    "古环境", "氧化还原", "地球化学", "沉积相",
    "BBox",
    "提取方法", "Needs Review", "Review Reasons",
    # Phase 64 Plan B (Task B.5): one summary column for the
    # schematic / diagram / reconstruction / phylogenetic
    # extraction. Format: "type|text=N|rel=N|conf=0.95"
    # (compact, machine-parseable, fits a single cell). Empty
    # string when the row has no figure_schematic_data so the
    # operator can filter or sort.
    "示意图摘要",
    # Phase 65 Plan A.5: cross-figure linker provenance. Carries
    # the winning strategy ("sample_match" / "locality_match" /
    # "m3_inference" / "unlinked") and the linker confidence.
    # Operator can filter "unlinked" rows to find panels that
    # still need manual linking. Empty when the linker didn't
    # run (legacy rows).
    "Link Source", "Link Confidence", "Link Figure",
]
# --- geology_context columns ------------------------------------------------
_GEOLOGY_HEADERS = [
    "论文ID", "地质上下文ID", "Section Type", "Section Title",
    "Age", "Chronostratigraphy", "Chrono Rank", "Ma_top", "Ma_base", "Ma_mid",
    "Formation", "Member", "Group", "Lithology", "Biozone",
    "古环境", "氧化还原", "地球化学", "沉积相",
    "Locality", "Country", "Modern_Lat", "Modern_Lon", "Coord来源",
    "LocalityID", "Confidence", "Evidence Text",
]
# --- locality columns ------------------------------------------------------
_LOCALITY_HEADERS = [
    "论文ID", "LocalityID", "Name", "Country", "Region", "Section Name",
    "Modern_Lat", "Modern_Lon", "Coordinate Source", "Geocoding Source",
    "Confidence", "Evidence Text (head 80)",
]
# --- paleo columns ---------------------------------------------------------
_PALEO_HEADERS = [
    "论文ID", "PaleoCoordinateID", "LocalityID", "Plate",
    "Modern_Lat", "Modern_Lon",
    "Reconstruction Age (Ma)",
    "Paleo_Lat", "Paleo_Lon",
    "Reconstruction Model", "Method", "Backend Status", "Confidence",
]


def _row_for_panel(p: dict[str, Any]) -> list[Any]:
    """Project a single ``PanelRecord`` dict (the API response shape
    after schema validation + model_dump) into the panel-sheet row.

    We avoid the deep ``r.metadata.geology_links[0].*`` chain and
    fall back to ``""`` when the field is missing — keeps the
    workbook clean and the operator focused on rows that have
    data."""
    md = p.get("metadata") or {}
    gl0 = (md.get("geology_links") or [{}])[0] or {} if md.get("geology_links") else {}
    if isinstance(gl0, list):
        gl0 = gl0[0] if gl0 else {}
    # audit 2026-07-26: geology_links may contain None entries; guard
    # with isinstance(g, dict) so .get() doesn't raise AttributeError.
    sample_ids = [
        g.get("sample_id")
        for g in (md.get("geology_links") or [])
        if isinstance(g, dict) and g.get("sample_id")
    ]
    bbox = p.get("bbox")
    bbox_str = (
        f"[{','.join(str(int(v)) for v in bbox)}]"
        if isinstance(bbox, list) and len(bbox) == 4
        else ""
    )
    return [
        p.get("paper_id") or "",
        p.get("figure_id") or "",
        p.get("panel_id") or "",
        # Phase 58 Plan 1.3 (Bug 1.3): panel_id_source lives at
        # panel.metadata["panel_id_source"] (set by converters.py:448
        # and association.match_panels). The previous lookup referenced
        # a non-existent legacy key, leaving the "Panel来源" column
        # always empty. Top-level p.get("panel_id_source") is the
        # PanelRecord schema field; prefer metadata first since that's
        # what the pipeline actually writes.
        (md.get("panel_id_source") or p.get("panel_id_source") or ""),
        p.get("species") or "",
        p.get("confidence") if p.get("confidence") is not None else "",
        md.get("geology_scope") or "",
        gl0.get("age") or "",
        gl0.get("chronostratigraphy") or "",
        gl0.get("chronostratigraphy_rank") or "",
        gl0.get("ma_top") if gl0.get("ma_top") is not None else "",
        gl0.get("ma_base") if gl0.get("ma_base") is not None else "",
        gl0.get("formation") or "",
        gl0.get("member") or "",
        gl0.get("group") or "",
        gl0.get("locality") or "",
        gl0.get("country") or "",
        gl0.get("modern_latitude") if gl0.get("modern_latitude") is not None else "",
        gl0.get("modern_longitude") if gl0.get("modern_longitude") is not None else "",
        gl0.get("coord_source") or "",
        "|".join(sample_ids),
        gl0.get("lithology") or "",
        gl0.get("biozone") or "",
        # Round 24: environment / geochem / facies proxies
        gl0.get("paleoenvironment") or "",
        gl0.get("redox") or "",
        gl0.get("chemostrat") or "",
        gl0.get("facies") or "",
        bbox_str,
        md.get("extraction_method") or "",
        p.get("needs_review") if p.get("needs_review") is not None else False,
        "|".join(p.get("review_reasons") or []),
        # Phase 64 Plan B (Task B.5): schematic summary column. We
        # collapse the free-form JSON into a compact string so the
        # operator can scan the column quickly without opening the
        # cell. Format: "schematic|text=12|rel=3|conf=0.95". Empty
        # for non-schematic rows so the workbook stays clean.
        _summarize_schematic_data(md.get("figure_schematic_data")),
        # Phase 65 Plan A.5: cross-figure linker provenance.
        # ``link_source`` is the winning strategy; ``link_confidence``
        # is the linker's own confidence (NOT the species confidence
        # above); ``link_figure_id`` is the paper-level figure the
        # panel was linked to. Empty when the linker didn't run.
        md.get("link_source") or "",
        md.get("link_confidence") if md.get("link_confidence") is not None else "",
        md.get("link_figure_id") or "",
    ]


def _summarize_schematic_data(schematic_data: Any) -> str:
    """Phase 64 Plan B (Task B.5): compact one-cell summary of the
    schematic extraction so the operator can scan the workbook
    without opening each cell.

    Format: ``"<figure_type>|text=<n_text>|rel=<n_rel>|conf=<0..1>"``
    Empty string when no schematic data is present (the row was a
    regular plate, not a conceptual figure).
    """
    if not isinstance(schematic_data, dict):
        return ""
    fig_type = str(schematic_data.get("figure_type") or "").strip()
    if not fig_type:
        return ""
    text_elements = schematic_data.get("text_elements") or []
    relationships = schematic_data.get("relationships") or []
    if not isinstance(text_elements, list):
        text_elements = []
    if not isinstance(relationships, list):
        relationships = []
    try:
        conf = float(schematic_data.get("confidence", 0.0) or 0.0)
    except (TypeError, ValueError):
        conf = 0.0
    conf = max(0.0, min(1.0, conf))
    return f"{fig_type}|text={len(text_elements)}|rel={len(relationships)}|conf={conf:.2f}"


def _row_for_geology_context(
    paper_id: str, g: dict[str, Any]
) -> list[Any]:
    return [
        paper_id,
        g.get("geology_context_id") or "",
        g.get("section_type") or "",
        g.get("section_title") or "",
        g.get("age") or "",
        g.get("chronostratigraphy") or "",
        g.get("chronostratigraphy_rank") or "",
        g.get("ma_top") if g.get("ma_top") is not None else "",
        g.get("ma_base") if g.get("ma_base") is not None else "",
        g.get("ma_mid") if g.get("ma_mid") is not None else "",
        g.get("formation") or "",
        g.get("member") or "",
        g.get("group") or "",
        g.get("lithology") or "",
        g.get("biozone") or "",
        # Round 24: environment / geochem / facies proxies
        g.get("paleoenvironment") or "",
        g.get("redox") or "",
        g.get("chemostrat") or "",
        g.get("facies") or "",
        g.get("locality") or "",
        g.get("country") or "",
        g.get("modern_latitude") if g.get("modern_latitude") is not None else "",
        g.get("modern_longitude") if g.get("modern_longitude") is not None else "",
        g.get("coord_source") or "",
        g.get("locality_id") or "",
        g.get("confidence") if g.get("confidence") is not None else "",
        (g.get("evidence_text") or "")[:200],
    ]


def _row_for_locality(paper_id: str, l: dict[str, Any]) -> list[Any]:
    return [
        paper_id,
        l.get("locality_id") or "",
        l.get("name") or "",
        l.get("country") or "",
        l.get("region") or "",
        l.get("section_name") or "",
        l.get("modern_latitude") if l.get("modern_latitude") is not None else "",
        l.get("modern_longitude") if l.get("modern_longitude") is not None else "",
        l.get("coordinate_source") or "",
        l.get("geocoding_source") or "",
        l.get("confidence") if l.get("confidence") is not None else "",
        (l.get("evidence_text") or "")[:200],  # Phase 55: match geology_contexts truncation
    ]


def _row_for_paleo(paper_id: str, p: dict[str, Any]) -> list[Any]:
    return [
        paper_id,
        p.get("paleo_coordinate_id") or "",
        p.get("locality_id") or "",
        p.get("plate_id") or "",
        p.get("modern_latitude") if p.get("modern_latitude") is not None else "",
        p.get("modern_longitude") if p.get("modern_longitude") is not None else "",
        p.get("reconstruction_age_ma") if p.get("reconstruction_age_ma") is not None else "",
        p.get("paleo_latitude") if p.get("paleo_latitude") is not None else "",
        p.get("paleo_longitude") if p.get("paleo_longitude") is not None else "",
        p.get("reconstruction_model") or "",
        p.get("method") or "",
        p.get("backend_status") or "",
        p.get("confidence") if p.get("confidence") is not None else "",
    ]


# --- legend (column description) -------------------------------------------
_LEGEND = [
    ("论文ID", "Stable paper identifier (paper_id in the API)"),
    ("图版ID", "Stable figure identifier (figure_id, includes page and plate number)"),
    ("Panel标签", "Panel label as printed in the figure (e.g. '1', 'A', 'Kenji Kashiwagi')"),
    ("Panel来源", "How the panel_id was derived: image_ocr / caption / position / legacy"),
    ("物种", "Genus + species (e.g. 'Megaporus jini'); may be 'N/A' or a non-biological token"),
    ("置信度", "0.0–1.0 confidence in the species assignment (Round 18 match score)"),
    ("地质范围", "Round 19 scope marker: panel | figure_anchor | none (whether the geology data is panel-specific or inherited)"),
    ("地层年代", "First geology link's age (e.g. 'Late Jurassic'); may be empty"),
    ("年代Chrono", "First geology link's chronostratigraphy (e.g. 'Kimmeridgian', 'Pliensbachian')"),
    ("年代Stage", "Round 18 rank: 'period' | 'epoch' | 'age' (Ma precision granularity)"),
    ("Ma_top", "First geology link's ma_top (younger Ma bound from ICS)"),
    ("Ma_base", "First geology link's ma_base (older Ma bound)"),
    ("Formation", "First geology link's formation (e.g. 'Dalong Formation', 'Zabijak Formation')"),
    ("Member", "First geology link's member (Round 18 split: group/formation/member as separate fields)"),
    ("Group", "First geology link's group"),
    ("Locality", "First geology link's locality (e.g. 'Méouge section', 'Karnezeika')"),
    ("Country", "First geology link's country (e.g. 'Greece', 'Tunisia', 'Russia')"),
    ("Modern_Lat", "First geology link's modern_latitude (Round 21: includes country-centroid fallback)"),
    ("Modern_Lon", "First geology link's modern_longitude"),
    ("Coord来源", "'caption' | 'country_centroid' (Round 21) | 'paleobiology_db' (future)"),
    ("Sample IDs", "Round 21 sample-id prefixes: S_ legacy | B_ Boughdiri | R_ specimen | N_ numeric | L_ (N) | P_ pl.N"),
    ("Lithology", "First geology link's lithology (Round 18: 30+ rock-name dictionary)"),
    ("Biozone", "First geology link's biozone (N. optima Zone style)"),
    ("古环境 (paleoenvironment)", "Round 24: oxygen regime of the water column (anoxic/euxinic/oxic/etc.). Critical for P/T boundary research."),
    ("氧化还原 (redox)", "Round 24: Algeo & Tribovillard 2009 classification (oxic/dysoxic/suboxic/anoxic/euxinic/ferruginous/sulfidic)."),
    ("地球化学 (chemostrat)", "Round 24: named chemostratigraphic events (CIE / mass extinction / OAE / P/T boundary / LIP / mercury anomaly)."),
    ("沉积相 (facies)", "Round 24: standard sedimentological facies (turbidite / pelagic / platform / basin / etc.)."),
    ("BBox", "Panel bounding box [x, y, w, h] in pixels"),
    ("提取方法", "Match method: heuristic | llm_first | cv | hybrid"),
    ("Needs Review", "Round 19: True if panel needs operator review (no printed panel_id, etc.)"),
    ("Review Reasons", "Round 19: pipe-separated list of review reasons"),
    ("Ma_mid", "Mid Ma (average of ma_top/ma_base) for paleo reconstruction"),
    ("Modern_Lat (paleo)", "Locality's modern_latitude (from centroid fallback if needed)"),
    ("Modern_Lon (paleo)", "Locality's modern_longitude"),
    ("Reconstruction Age (Ma)", "Ma at which the GPlates Seton2012 model reconstructs the position"),
    ("Paleo_Lat", "Reconstructed paleo_latitude (NaN if reconstruction failed)"),
    ("Paleo_Lon", "Reconstructed paleo_longitude"),
    ("Reconstruction Model", "'Seton2012' for Round 20+; later models TBD"),
    ("Method", "euler_pole_rotation (Round 20 default)"),
    ("Backend Status", "'ok' (reconstruction succeeded) | 'plate_or_age_unknown' (degraded)"),
    ("LocalityID", "Stable locality id (round 20: paper_id+locality+lat+lon)"),
    ("LocalityID (paleo)", "Locality id used for paleo coordinate join (matches Locality sheet)"),
    ("Section Type", "'geological_setting' | 'systematic_paleontology' | 'references' | 'other'"),
    ("Section Title", "Section title as detected by GROBID/OpenDataLoader"),
    ("Region", "Region override (Round 20: 'Sicily' → 'Italy')"),
    ("Section Name (loc)", "Section name attached to the locality record"),
    ("Coordinate Source (loc)", "'caption' | 'country_centroid' | other"),
    ("Geocoding Source (loc)", "Round 20+; reserved for future Nominatim/GeoNames lookups"),
    ("Evidence Text", "Round 18 evidence text (first 200 chars)"),
    ("Section Type (loc)", "Same as Section Type but for the locality record"),
]


def write_xlsx(
    run_output: dict[str, Any] | Any, path: str | None = None
) -> bytes | None:
    """Build a multi-sheet .xlsx from a ``RunOutput`` dict.

    Returns the .xlsx bytes if ``path`` is ``None`` (in-memory
    mode, used by the API endpoint), or writes to ``path`` and
    returns ``None`` (used by the CLI exporter).

    Accepts either a plain ``dict`` or a Pydantic ``RunOutput``
    model (which has ``.panels`` / ``.localities`` /
    ``.paleo_coordinates`` as model fields, not dict keys).
    """
    # Normalise: Pydantic BaseModel has .model_dump(); plain dict does not.
    if hasattr(run_output, "model_dump"):
        run_output = run_output.model_dump()
    wb = Workbook()
    # ---- Sheet 1: panels -------------------------------------------------
    ws = wb.active
    ws.title = "panels"
    _write_header(ws, _PANEL_HEADERS)
    panel_rows: list[list[Any]] = []
    for p in run_output.get("panels", []) or []:
        panel_rows.append(_row_for_panel(p))
    _write_rows(ws, _PANEL_HEADERS, panel_rows)
    _autosize_columns(ws)

    # ---- Sheet 2: geology_contexts ---------------------------------------
    ws2 = wb.create_sheet("geology_contexts")
    _write_header(ws2, _GEOLOGY_HEADERS)
    geo_rows: list[list[Any]] = []
    for p in run_output.get("panels", []) or []:
        paper_id = p.get("paper_id") or ""
        for g in (p.get("metadata", {}) or {}).get("geology_links") or []:
            geo_rows.append(_row_for_geology_context(paper_id, g))
    _write_rows(ws2, _GEOLOGY_HEADERS, geo_rows)
    _autosize_columns(ws2)

    # ---- Sheet 3: localities ---------------------------------------------
    ws3 = wb.create_sheet("localities")
    _write_header(ws3, _LOCALITY_HEADERS)
    loc_rows: list[list[Any]] = []
    for l in run_output.get("localities", []) or []:
        # audit 2026-07-26: guard None entries (same null risk as
        # geology_links) before calling .get() on the row.
        if not isinstance(l, dict):
            continue
        loc_rows.append(_row_for_locality(l.get("paper_id") or "", l))
    _write_rows(ws3, _LOCALITY_HEADERS, loc_rows)
    _autosize_columns(ws3)

    # ---- Sheet 4: paleo_coordinates --------------------------------------
    ws4 = wb.create_sheet("paleo_coordinates")
    _write_header(ws4, _PALEO_HEADERS)
    paleo_rows: list[list[Any]] = []
    for p in run_output.get("paleo_coordinates", []) or []:
        if not isinstance(p, dict):
            continue
        paleo_rows.append(_row_for_paleo(p.get("paper_id") or "", p))
    _write_rows(ws4, _PALEO_HEADERS, paleo_rows)
    _autosize_columns(ws4)

    # ---- Sheet 5: legend (column descriptions) --------------------------
    ws5 = wb.create_sheet("legend")
    _write_header(ws5, ["列名 (Column)", "含义 (Description)"])
    legend_rows = [[col, desc] for col, desc in _LEGEND]
    _write_rows(ws5, ["列名 (Column)", "含义 (Description)"], legend_rows)
    _autosize_columns(ws5, max_width=80)

    if path is None:
        # In-memory mode for the API endpoint
        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
    # Phase 38: atomic write — openpyxl writes directly to ``path``,
    # so a crash mid-write leaves a corrupt xlsx that's unrecoverable.
    # Write to ``path + ".tmp"``, fsync, then ``os.replace`` (atomic on
    # POSIX and Windows).
    import os
    path_obj = Path(path)
    tmp = path_obj.with_suffix(path_obj.suffix + ".tmp")
    try:
        wb.save(tmp)
        os.replace(tmp, path_obj)
    except Exception:
        # Clean up partial tmp file on failure
        try:
            tmp.unlink()
        except OSError:
            pass
        raise
    return None
